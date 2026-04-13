"""openpyxl workbook builder for teacher export.

Produces three sheets:
  • Raw    — one row per latest submission
  • Matrix — target × grader → latest total
  • Summary — target × period → average / median / max / min / count
"""
from __future__ import annotations

import sqlite3
from statistics import mean, median
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font


RAW_HEADERS = [
    "評分者學號", "評分者姓名", "評分者班級",
    "被評者學號", "被評者姓名", "被評者班級",
    "場次",
    "主題掌握", "內容豐富", "敘事技巧", "簡報技巧與互動", "團隊表現",
    "總分", "建議或留言", "提交時間",
]


def _period_label(conn: sqlite3.Connection, code: str) -> str:
    row = conn.execute(
        "SELECT label FROM evaluation_periods WHERE code = ?", (code,)
    ).fetchone()
    return row["label"] if row else code


def build_workbook(conn: sqlite3.Connection, period_code: str) -> Workbook:
    label = _period_label(conn, period_code)
    rows = conn.execute(
        "SELECT ls.*, gu.name AS grader_name, gu.class_name AS grader_class,"
        "       tu.name AS target_name, tu.class_name AS target_class "
        "FROM latest_submissions ls "
        "JOIN permitted_users gu ON gu.student_id = ls.grader_student_id "
        "JOIN permitted_users tu ON tu.student_id = ls.target_student_id "
        "WHERE ls.period_code = ?",
        (period_code,),
    ).fetchall()
    students = conn.execute(
        "SELECT student_id, name, class_name FROM permitted_users ORDER BY student_id"
    ).fetchall()

    wb = Workbook()

    # Sheet 1 — Raw
    raw = wb.active
    raw.title = "Raw"
    raw.append(RAW_HEADERS)
    for c in raw[1]:
        c.font = Font(bold=True)
    for r in rows:
        total = (
            r["score_topic"] + r["score_content"] + r["score_narrative"]
            + r["score_presentation"] + r["score_teamwork"]
        )
        raw.append([
            r["grader_student_id"], r["grader_name"], r["grader_class"],
            r["target_student_id"], r["target_name"], r["target_class"],
            label,
            r["score_topic"], r["score_content"], r["score_narrative"],
            r["score_presentation"], r["score_teamwork"],
            total, r["comment"], r["submitted_at"],
        ])

    # Sheet 2 — Matrix
    matrix = wb.create_sheet("Matrix")
    header = [""] + [f"{s['student_id']} {s['name']}" for s in students]
    matrix.append(header)
    for c in matrix[1]:
        c.font = Font(bold=True)

    totals_map: Dict[str, Dict[str, int]] = {}
    for r in rows:
        total = (
            r["score_topic"] + r["score_content"] + r["score_narrative"]
            + r["score_presentation"] + r["score_teamwork"]
        )
        totals_map.setdefault(r["target_student_id"], {})[r["grader_student_id"]] = total

    for target in students:
        row_values: List[Any] = [f"{target['student_id']} {target['name']}"]
        for grader in students:
            if grader["student_id"] == target["student_id"]:
                row_values.append("")
            else:
                row_values.append(
                    totals_map.get(target["student_id"], {}).get(grader["student_id"], "")
                )
        matrix.append(row_values)

    # Sheet 3 — Summary
    summary = wb.create_sheet("Summary")
    summary.append(["學號", "姓名", "場次", "平均", "中位", "最大", "最小", "受評次數"])
    for c in summary[1]:
        c.font = Font(bold=True)
    for target in students:
        totals = [
            (
                r["score_topic"] + r["score_content"] + r["score_narrative"]
                + r["score_presentation"] + r["score_teamwork"]
            )
            for r in rows
            if r["target_student_id"] == target["student_id"]
        ]
        if totals:
            summary.append([
                target["student_id"], target["name"], label,
                round(mean(totals), 2), median(totals), max(totals), min(totals), len(totals),
            ])
        else:
            summary.append([target["student_id"], target["name"], label, "", "", "", "", 0])

    return wb
