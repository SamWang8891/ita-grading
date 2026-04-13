from __future__ import annotations

import io

from openpyxl import load_workbook

from tests.conftest import login_student, login_teacher, seed


def _submit(client, grader, target, scores, comment=""):
    with client:
        client.post("/api/auth/logout")
    login_student(client, grader)
    return client.post("/api/student/submissions", json={
        "period": "midterm", "target_student_id": target,
        "scores": scores, "comment": comment, "self_note": "hidden",
    })


def _grade(client, grader, target, topic, content, narrative, presentation, teamwork, comment=""):
    client.post("/api/auth/logout")
    login_student(client, grader)
    r = client.post("/api/student/submissions", json={
        "period": "midterm", "target_student_id": target,
        "scores": {"topic": topic, "content": content, "narrative": narrative,
                   "presentation": presentation, "teamwork": teamwork},
        "comment": comment, "self_note": "private",
    })
    assert r.status_code == 200, r.text


def test_overview_statistics(app_client):
    seed(app_client)
    _grade(app_client, "B001", "B002", 30, 30, 20, 10, 10, comment="A")
    _grade(app_client, "B003", "B002", 20, 20, 10, 5, 5, comment="B")

    app_client.post("/api/auth/logout")
    login_teacher(app_client, "t1", "tpw")
    overview = app_client.get("/api/teacher/overview?period=midterm").json()
    b002 = next(s for s in overview if s["student_id"] == "B002")
    assert b002["stats"]["count"] == 2
    assert b002["stats"]["max"] == 100
    assert b002["stats"]["min"] == 60
    assert b002["stats"]["median"] == 80
    # self_note must never be exposed to the teacher
    for row in b002["received"]:
        assert "self_note" not in row


def test_excel_export_three_sheets(app_client):
    seed(app_client)
    _grade(app_client, "B001", "B002", 20, 20, 10, 5, 5)
    _grade(app_client, "B003", "B002", 30, 20, 10, 5, 5)
    app_client.post("/api/auth/logout")
    login_teacher(app_client, "t1", "tpw")
    r = app_client.get("/api/teacher/export.xlsx?period=midterm")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Raw", "Matrix", "Summary"]
    raw_header = [c.value for c in wb["Raw"][1]]
    assert raw_header[0] == "評分者學號"
    assert raw_header[-1] == "提交時間"
    # Matrix: first row = headers; corner empty (openpyxl stores "" as None); has student columns
    matrix_header = [c.value for c in wb["Matrix"][1]]
    assert matrix_header[0] in (None, "")
    assert any("B002" in str(c) for c in matrix_header[1:])
    # Summary: target × period row
    assert wb["Summary"][1][0].value == "學號"


def test_student_cannot_hit_teacher(app_client):
    seed(app_client)
    login_student(app_client, "B001")
    assert app_client.get("/api/teacher/overview?period=midterm").status_code == 403


def test_teacher_change_password(app_client):
    seed(app_client)
    login_teacher(app_client, "t1", "tpw")
    ok = app_client.post("/api/teacher/password",
                         json={"old_password": "tpw", "new_password": "new-pass"})
    assert ok.status_code == 200
    # old password now fails
    app_client.post("/api/auth/logout")
    r = app_client.post("/api/auth/password", json={"identifier": "t1", "password": "tpw"})
    assert r.status_code == 401
    ok2 = app_client.post("/api/auth/password", json={"identifier": "t1", "password": "new-pass"})
    assert ok2.status_code == 200
